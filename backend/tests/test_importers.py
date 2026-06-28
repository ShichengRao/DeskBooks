from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app import importers as importer_registry
from app import schemas
from app.importers.amex import AmexImporter
from app.importers.amex_xlsx import parse_amex_xlsx_bytes
from app.importers.chase_credit import ChaseCreditImporter
from app.importers.contribution_history import ContributionHistoryImporter
from app.importers.us_banks import (
    CapitalOneCreditImporter,
    ChaseBankImporter,
    CitiCreditImporter,
    DebitCreditBankImporter,
    DiscoverCreditImporter,
    MarcusMorganStanleyBankImporter,
    PncBankImporter,
    RunningBalanceBankImporter,
    UsBankImporter,
)
from app.importers.wells_fargo_checking import WellsFargoCheckingImporter
from app.models import (
    Account,
    AccountCategory,
    AccountType,
    Base,
    NetWorthSnapshot,
    SignConvention,
    TransactionKind,
)
from app.routers import imports, snapshots


def test_chase_credit_importer_marks_payments_and_returns():
    rows = ChaseCreditImporter.parse(
        "\n".join(
            [
                "Transaction Date,Post Date,Description,Category,Type,Amount,Memo",
                "06/01/2026,06/02/2026,ACME GROCERY,Groceries,Sale,-42.18,",
                "06/10/2026,06/10/2026,AUTOPAY PAYMENT,Payment,Payment,50.68,",
                "06/12/2026,06/13/2026,RETURN MERCHANT,Shopping,Return,12.00,",
            ]
        )
    )

    assert len(rows) == 3
    assert rows[0].amount == Decimal("-42.18")
    assert rows[0].suggested_kind == TransactionKind.uncategorized
    assert rows[1].suggested_kind == TransactionKind.cc_payment
    assert rows[2].suggested_kind == TransactionKind.refund


def test_amex_importer_flips_charges_positive_exports():
    rows = AmexImporter.parse(
        "\n".join(
            [
                "Date,Description,Amount",
                "06/01/2026,ONLINE SOFTWARE SUBSCRIPTION,19.00",
                "06/09/2026,PAYMENT RECEIVED,-33.25",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("-19.00"), Decimal("33.25")]
    assert rows[1].suggested_kind == TransactionKind.cc_payment


def test_amex_xlsx_importer_reads_transaction_details_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction Details"
    ws.append(["Downloaded", "Example"])
    ws.append(["Date", "Description", "Amount"])
    ws.append(["06/01/2026", "ONLINE SOFTWARE SUBSCRIPTION", 19.00])
    ws.append(["06/09/2026", "PAYMENT RECEIVED", -33.25])
    buf = BytesIO()
    wb.save(buf)

    rows = parse_amex_xlsx_bytes(buf.getvalue())

    assert [r.amount for r in rows] == [Decimal("-19.00"), Decimal("33.25")]
    assert rows[1].suggested_kind == TransactionKind.cc_payment


def test_wells_fargo_importer_marks_common_local_kinds():
    rows = WellsFargoCheckingImporter.parse(
        "\n".join(
            [
                "DATE,DESCRIPTION,AMOUNT,CHECK #,STATUS",
                "06/01/2026,ACME PAYROLL,2500.00,,POSTED",
                "06/05/2026,CHASE CREDIT CRD AUTOPAY,-250.00,,POSTED",
                "06/08/2026,NEIGHBORHOOD COFFEE,-6.75,,POSTED",
            ]
        )
    )

    assert [r.suggested_kind for r in rows] == [
        TransactionKind.income,
        TransactionKind.cc_payment,
        TransactionKind.uncategorized,
    ]


def test_chase_bank_importer_reads_signed_bank_activity():
    rows = ChaseBankImporter.parse(
        "\n".join(
            [
                "Details,Posting Date,Description,Amount,Type,Balance,Check or Slip #",
                "DEBIT,06/01/2026,NEIGHBORHOOD COFFEE,-6.75,DEBIT_CARD,993.25,",
                "CREDIT,06/02/2026,ACME PAYROLL,2500.00,ACH_CREDIT,3493.25,",
                "DEBIT,06/05/2026,ONLINE TRANSFER TO SAVINGS,-300.00,ACH_DEBIT,3193.25,",
            ]
        )
    )

    assert [r.amount for r in rows] == [
        Decimal("-6.75"),
        Decimal("2500.00"),
        Decimal("-300.00"),
    ]
    assert rows[1].suggested_kind == TransactionKind.income
    assert rows[2].suggested_kind == TransactionKind.transfer


def test_running_balance_bank_importer_covers_boa_truist_bmo_shape():
    rows = RunningBalanceBankImporter.parse(
        "\n".join(
            [
                "Date,Description,Amount,Running Bal.",
                "06/01/2026,POINT OF SALE PURCHASE,-42.18,1250.00",
                "06/03/2026,TREAS 310 TAX REF,150.00,1400.00",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("-42.18"), Decimal("150.00")]
    assert rows[1].suggested_kind == TransactionKind.income


def test_pnc_bank_importer_converts_withdrawals_and_deposits():
    rows = PncBankImporter.parse(
        "\n".join(
            [
                "Date,Description,Withdrawals,Deposits,Balance",
                "06/01/2026,ATM WITHDRAWAL,80.00,,920.00",
                "06/02/2026,DIRECT DEP PAYROLL,,2500.00,3420.00",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("-80.00"), Decimal("2500.00")]
    assert rows[1].suggested_kind == TransactionKind.income


def test_debit_credit_bank_importer_covers_td_and_custody_shape():
    rows = DebitCreditBankImporter.parse(
        "\n".join(
            [
                "Date,Description,Debit,Credit,Balance",
                "06/01/2026,ACH TRANSFER TO BROKERAGE,500.00,,1500.00",
                "06/04/2026,INTEREST PAID,,1.23,1501.23",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("-500.00"), Decimal("1.23")]
    assert rows[0].suggested_kind == TransactionKind.transfer
    assert rows[1].suggested_kind == TransactionKind.income


def test_us_bank_importer_combines_name_transaction_and_memo():
    rows = UsBankImporter.parse(
        "\n".join(
            [
                "Date,Transaction,Name,Memo,Amount",
                "06/01/2026,DEBIT_CARD,NEIGHBORHOOD COFFEE,Latte,-6.75",
                "06/02/2026,ACH_CREDIT,ACME PAYROLL,Salary,2500.00",
            ]
        )
    )

    assert rows[0].description_normalized == "NEIGHBORHOOD COFFEE DEBIT_CARD Latte"
    assert [r.amount for r in rows] == [Decimal("-6.75"), Decimal("2500.00")]
    assert rows[1].suggested_kind == TransactionKind.income


def test_activity_bank_importer_covers_marcus_and_morgan_stanley_shape():
    rows = MarcusMorganStanleyBankImporter.parse(
        "\n".join(
            [
                "Date,Activity,Description,Amount,Balance",
                "06/01/2026,Interest,Monthly interest,3.21,1003.21",
                "06/02/2026,Transfer,External transfer,-100.00,903.21",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("3.21"), Decimal("-100.00")]
    assert rows[0].suggested_kind == TransactionKind.income
    assert rows[1].suggested_kind == TransactionKind.transfer


def test_capital_one_credit_importer_converts_debit_credit_columns():
    rows = CapitalOneCreditImporter.parse(
        "\n".join(
            [
                "Transaction Date,Posted Date,Card No.,Description,Category,Debit,Credit",
                "06/01/2026,06/02/2026,1234,ONLINE SOFTWARE,Services,19.00,",
                "06/10/2026,06/10/2026,1234,AUTOPAY PAYMENT,Payment,,50.00",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("-19.00"), Decimal("50.00")]
    assert rows[1].suggested_kind == TransactionKind.cc_payment


def test_citi_credit_importer_converts_debit_credit_columns():
    rows = CitiCreditImporter.parse(
        "\n".join(
            [
                "Status,Date,Description,Debit,Credit",
                "Cleared,06/01/2026,BOOKSTORE,31.42,",
                "Cleared,06/12/2026,PAYMENT THANK YOU,,31.42",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("-31.42"), Decimal("31.42")]
    assert rows[1].suggested_kind == TransactionKind.cc_payment


def test_discover_credit_importer_flips_charges_positive_exports():
    rows = DiscoverCreditImporter.parse(
        "\n".join(
            [
                "Trans. Date,Post Date,Description,Amount,Category",
                "06/01/2026,06/02/2026,RESTAURANT,23.45,Restaurants",
                "06/10/2026,06/10/2026,PAYMENT RECEIVED,-23.45,Payments",
            ]
        )
    )

    assert [r.amount for r in rows] == [Decimal("-23.45"), Decimal("23.45")]
    assert rows[1].suggested_kind == TransactionKind.cc_payment


def test_us_bank_importer_sniffing_prefers_specific_over_generic_shapes():
    citi_csv = "\n".join(
        [
            "Status,Date,Description,Debit,Credit",
            "Cleared,06/01/2026,BOOKSTORE,31.42,",
        ]
    )
    activity_csv = "\n".join(
        [
            "Date,Activity,Description,Amount,Balance",
            "06/01/2026,Interest,Monthly interest,3.21,1003.21",
        ]
    )

    assert importer_registry.sniff(citi_csv)[0] is CitiCreditImporter
    assert importer_registry.sniff(activity_csv)[0] is MarcusMorganStanleyBankImporter


def test_importer_list_exposes_one_amex_choice():
    rows = imports.list_importers()
    amex_rows = [row for row in rows if "amex" in row["name"]]

    assert amex_rows == [{"name": "amex", "label": "Amex"}]


def test_contribution_history_importer_skips_metadata_rows():
    rows = ContributionHistoryImporter.parse(
        "\n".join(
            [
                "CONTRIBUTION HISTORY",
                "",
                "Account,Example Fund",
                "",
                'Status,Description,Symbol,"Estimated Amount","Received Date","Contribution ID"',
                'Complete,"INDEX ETF",ETF,"$123.45","2026-05-26T00:00:00-04:00",abc-1',
            ]
        )
    )

    assert len(rows) == 1
    assert rows[0].date.isoformat() == "2026-05-26"
    assert rows[0].amount == Decimal("-123.45")
    assert rows[0].suggested_kind == TransactionKind.donation


def test_net_worth_workbook_import_uses_user_supplied_account_map(tmp_path):
    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, future=True)
    db = Session()
    db.add(
        Account(
            name="Checking",
            account_category=AccountCategory.bank,
            type=AccountType.checking,
            sign_convention=SignConvention.outflow_negative,
        )
    )
    db.commit()

    wb = openpyxl.Workbook()
    dates = wb.active
    dates.title = "Dates"
    dates.cell(1, 2).value = date(2026, 6, 1)
    assets = wb.create_sheet("Assets")
    assets.cell(2, 1).value = "Bank row from arbitrary workbook"
    assets.cell(2, 2).value = 1234.56
    workbook_path = tmp_path / "net-worth.xlsx"
    wb.save(workbook_path)

    result = snapshots.import_workbook(
        schemas.NetWorthWorkbookImportRequest(
            path=str(workbook_path),
            account_map={"Assets!2": "Checking"},
        ),
        db,
    )

    assert result.imported == 1
    snap = db.scalar(select(NetWorthSnapshot))
    assert snap is not None
    assert snap.snapshot_date == date(2026, 6, 1)
    assert snap.balances[0].balance == Decimal("1234.56")
