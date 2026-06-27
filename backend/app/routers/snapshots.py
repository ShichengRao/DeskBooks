from __future__ import annotations

from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import APIRouter, HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from .. import analytics, models, schemas
from .common import DbSession, get_or_404

router = APIRouter(prefix="/api/snapshots", tags=["snapshots"])


@router.get("", response_model=list[schemas.NetWorthSnapshotOut])
def list_snapshots(db: DbSession):
    stmt = (
        select(models.NetWorthSnapshot)
        .options(selectinload(models.NetWorthSnapshot.balances))
        .order_by(models.NetWorthSnapshot.snapshot_date.desc())
    )
    return list(db.scalars(stmt))


@router.post("", response_model=schemas.NetWorthSnapshotOut)
def create_snapshot(body: schemas.NetWorthSnapshotIn, db: DbSession):
    existing = db.scalar(
        select(models.NetWorthSnapshot).where(models.NetWorthSnapshot.snapshot_date == body.snapshot_date)
    )
    if existing:
        raise HTTPException(409, "snapshot for this date already exists")
    snap = models.NetWorthSnapshot(snapshot_date=body.snapshot_date, notes=body.notes)
    db.add(snap)
    db.flush()
    for b in body.balances:
        db.add(
            models.AccountBalance(
                snapshot_id=snap.id, account_id=b.account_id, balance=b.balance, notes=b.notes
            )
        )
    db.commit()
    db.refresh(snap)
    return snap


@router.post("/import-workbook", response_model=schemas.NetWorthWorkbookImportResult)
def import_workbook(body: schemas.NetWorthWorkbookImportRequest, db: DbSession):
    path = Path(body.path).expanduser()
    if not path.exists() or not path.is_file():
        raise HTTPException(404, "file not found")
    wb = load_workbook(path, data_only=True, read_only=False)
    if "Dates" not in wb.sheetnames:
        raise HTTPException(400, "workbook is missing a Dates sheet")

    account_by_name = {account.name: account for account in db.scalars(select(models.Account)).all()}
    workbook_rows, missing = _mapped_workbook_rows(wb, body.account_map, account_by_name)
    if missing:
        return schemas.NetWorthWorkbookImportResult(
            imported=0,
            skipped_existing=0,
            missing_accounts=missing,
        )

    dates_ws = wb["Dates"]
    existing_dates = set(db.scalars(select(models.NetWorthSnapshot.snapshot_date)).all())
    imported = 0
    skipped = 0
    for col in range(2, dates_ws.max_column + 1):
        raw_date = dates_ws.cell(1, col).value
        if raw_date is None:
            continue
        snap_date = raw_date.date() if hasattr(raw_date, "date") else raw_date
        if snap_date in existing_dates:
            skipped += 1
            continue
        balances = []
        for sheet_name, row, account_name in workbook_rows:
            value = _decimal_or_none(wb[sheet_name].cell(row, col).value)
            if value is None:
                continue
            balances.append(
                models.AccountBalance(
                    account_id=account_by_name[account_name].id,
                    balance=value,
                )
            )
        snap = models.NetWorthSnapshot(
            snapshot_date=snap_date,
            notes=f"Imported from {path.name}",
            balances=balances,
        )
        db.add(snap)
        imported += 1
    db.commit()
    return schemas.NetWorthWorkbookImportResult(
        imported=imported,
        skipped_existing=skipped,
        missing_accounts=[],
    )


def _mapped_workbook_rows(wb, account_map: dict[str, str], account_by_name: dict[str, models.Account]):
    if account_map:
        rows = []
        missing_accounts = set()
        invalid_keys = []
        for key, account_name in account_map.items():
            if account_name not in account_by_name:
                missing_accounts.add(account_name)
                continue
            if "!" not in key:
                invalid_keys.append(key)
                continue
            sheet_name, row_text = key.rsplit("!", 1)
            if sheet_name not in wb.sheetnames:
                invalid_keys.append(key)
                continue
            try:
                row = int(row_text)
            except ValueError:
                invalid_keys.append(key)
                continue
            rows.append((sheet_name, row, account_name))
        if invalid_keys:
            raise HTTPException(400, f"invalid account_map row key(s): {', '.join(sorted(invalid_keys))}")
        return rows, sorted(missing_accounts)

    rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "Dates":
            continue
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row, 1).value
            if isinstance(label, str) and label.strip() in account_by_name:
                rows.append((sheet_name, row, label.strip()))
    if not rows:
        raise HTTPException(
            400,
            "No account rows were found. Add a JSON account map using keys like \"Sheet name!12\" and account names as values.",
        )
    return rows, []


def _decimal_or_none(value) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


@router.patch("/{snap_id}", response_model=schemas.NetWorthSnapshotOut)
def update_snapshot(snap_id: int, body: schemas.NetWorthSnapshotUpdate, db: DbSession):
    snap = get_or_404(db, models.NetWorthSnapshot, snap_id)
    if body.snapshot_date is not None:
        snap.snapshot_date = body.snapshot_date
    if body.notes is not None:
        snap.notes = body.notes
    if body.balances is not None:
        # upsert balances
        existing = {b.account_id: b for b in snap.balances}
        seen = set()
        for b in body.balances:
            seen.add(b.account_id)
            if b.account_id in existing:
                existing[b.account_id].balance = b.balance
                existing[b.account_id].notes = b.notes
            else:
                db.add(
                    models.AccountBalance(
                        snapshot_id=snap.id,
                        account_id=b.account_id,
                        balance=b.balance,
                        notes=b.notes,
                    )
                )
        # delete balances not in payload
        for acc_id, bal in existing.items():
            if acc_id not in seen:
                db.delete(bal)
    db.commit()
    db.refresh(snap)
    return snap


@router.delete("/{snap_id}")
def delete_snapshot(snap_id: int, db: DbSession):
    snap = get_or_404(db, models.NetWorthSnapshot, snap_id)
    db.delete(snap)
    db.commit()
    return {"status": "deleted"}


@router.get("/series")
def snapshot_series(
    db: DbSession,
    start: date | None = None,
    end: date | None = None,
):
    if start is not None and end is not None and end < start:
        raise HTTPException(400, "end must be on or after start")
    return analytics.networth_series(db, start=start, end=end)
