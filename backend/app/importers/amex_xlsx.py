"""Amex xlsx exports.

Amex switched their consumer export to .xlsx with metadata rows before
the actual header. Sheet "Transaction Details" usually has several metadata
rows first, then a transaction header row where column A is "Date".

Charges are still positive — we invert to outflow-negative on the way in
to keep the rest of the app uniform.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

import openpyxl

from ..models import TransactionKind
from ..schemas import ImportDraftRow
from .base import guess_merchant, normalize_description


def parse_amex_xlsx_path(path: Path) -> list[ImportDraftRow]:
    return parse_amex_xlsx_bytes(path.read_bytes())


def parse_amex_xlsx_bytes(data: bytes) -> list[ImportDraftRow]:
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    if "Transaction Details" in wb.sheetnames:
        ws = wb["Transaction Details"]
    else:
        ws = wb[wb.sheetnames[0]]

    # Locate the header row: the first row whose col A == "Date".
    header_row = None
    for r in range(1, min(ws.max_row + 1, 30)):
        if (ws.cell(row=r, column=1).value or "").__str__().strip().lower() == "date":
            header_row = r
            break
    if header_row is None:
        return []

    out: list[ImportDraftRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        d_raw = ws.cell(row=r, column=1).value
        desc_raw = ws.cell(row=r, column=2).value
        amt_raw = ws.cell(row=r, column=3).value
        if d_raw is None or desc_raw is None or amt_raw is None:
            continue
        # Coerce date — can be a datetime, date, or "MM/DD/YYYY" string.
        if hasattr(d_raw, "date"):
            d = d_raw.date()
        elif hasattr(d_raw, "year"):
            d = d_raw  # already a date
        else:
            from dateutil import parser as dp

            try:
                d = dp.parse(str(d_raw)).date()
            except Exception:
                continue
        try:
            amount = Decimal(str(amt_raw)).quantize(Decimal("0.01"))
        except Exception:
            continue
        # Charges are positive in Amex's format; flip to outflow-negative.
        amount = -amount
        desc = normalize_description(str(desc_raw))
        upper = desc.upper()
        kind = TransactionKind.uncategorized
        if "AUTOPAY PAYMENT" in upper or "PAYMENT - THANK YOU" in upper or "PAYMENT RECEIVED" in upper:
            kind = TransactionKind.cc_payment
        out.append(
            ImportDraftRow(
                row_index=r - header_row - 1,
                date=d,
                post_date=None,
                description_raw=str(desc_raw),
                description_normalized=desc,
                merchant=guess_merchant(desc),
                amount=amount,
                suggested_kind=kind,
                raw={"row": r},
            )
        )
    return out
